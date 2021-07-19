from fuzzywuzzy import process, fuzz
import sys
from openpyxl import Workbook
from openpyxl import load_workbook


def excel():
    filename = sys.argv[1]
    filename2 = sys.argv[2]

    std_workbook = load_workbook(filename=filename, read_only=True)
    workbook = load_workbook(filename=filename2, read_only=True)

    std_sheet = std_workbook.active

    workbook_result = Workbook()
    sheet_result = workbook_result.active

    sheet_result["A1"] = "hello"
    sheet_result["B1"] = "world!"

    # Using the values_only because you just want to return the cell value
    for row in sheet_result.iter_rows(min_row=2, values_only=True):
        print(row)

    workbook_result.save(filename="combined_result.xlsx")


sample_words = ["date of birth", "birth_date", "birth date", "ayush suman", "date of join"]
reference = ["ayush", "joining date", "birth day"]


# Input Arrays :
# sample words : ["date of birth", "birth_date", "birth date", "ayush suman", "date of join"]
# reference words or the words that are considered to be standard : ["ayush", "joining date", "birth day"]

# algorithm to identify closest match
def loop_over_for_max_score():
    closest_match_dict = {}
    for i in sample_words:
        max_score = 0
        closest_match = ""
        for j in reference:
            score = fuzz.token_sort_ratio(i, j)
            # print(i, j, score)
            if max_score < score:
                max_score = score
                closest_match = j
        # print(i, "matches closest with", closest_match)
        closest_match_dict[closest_match] = i
    return closest_match_dict


# Output :
# date of birth ayush 11
# date of birth joining date 48
# date of birth birth day 73
# date of birth matches closest with birth day
# birth_date ayush 13
# birth_date joining date 36
# birth_date birth day 74
# birth_date matches closest with birth day
# birth date ayush 13
# birth date joining date 36
# birth date birth day 84
# birth date matches closest with birth day
# ayush suman ayush 62
# ayush suman joining date 26
# ayush suman birth day 30
# ayush suman matches closest with ayush
# date of join ayush 12
# date of join joining date 75
# date of join birth day 19
# date of join matches closest with joining date


# algorithm using process extractOne
def extract_using_process():
    closest_match_dict = {}
    for i in sample_words:
        closest_match = process.extractOne(i, reference, scorer=fuzz.token_sort_ratio, score_cutoff=50)
        closest_match_dict[closest_match.__getitem__(0)] = i
        # print(i, "matches closes with", closest_match.__getitem__(0), closest_match.__getitem__(1))
    return closest_match_dict


# Output:
# date of birth matches closes with birth day 73
# birth_date matches closes with birth day 74
# birth date matches closes with birth day 84
# ayush suman matches closes with ayush 62
# date of join matches closes with joining date 75

if __name__ == "__main__":
    closest = extract_using_process()
    print(closest)
    closest = loop_over_for_max_score()
    print(closest)
    excel()
